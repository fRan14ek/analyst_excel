from __future__ import annotations

import contextlib
import datetime as dt
import io
import itertools
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional

import pandas as pd
import typer
import yaml
from loguru import logger
from openpyxl import load_workbook
from pydantic import BaseModel, validator


class PathConfig(BaseModel):
    data_dir: Path
    base_file: Path
    input_dir: Path
    output_dir: Path
    logs_dir: Path
    lookup_product: Path
    columns_registry: Path

    @validator("data_dir", "base_file", "input_dir", "output_dir", "logs_dir", "lookup_product", "columns_registry")
    def _resolve_path(cls, value: Path) -> Path:
        return Path(value)


class MappingConfig(BaseModel):
    core: Path
    aliases: Mapping[str, Path]

    @validator("core")
    def _resolve_core(cls, value: Path) -> Path:
        return Path(value)

    @validator("aliases", pre=True)
    def _resolve_aliases(cls, value: Mapping[str, Path]) -> Mapping[str, Path]:
        return {k: Path(v) for k, v in value.items()}


class ProcessingConfig(BaseModel):
    enable_parquet: bool = True
    default_platforms: List[str]
    id_column: str = "id_key"


class AppConfig(BaseModel):
    paths: PathConfig
    mappings: MappingConfig
    processing: ProcessingConfig

    def resolve(self, root: Path) -> "AppConfig":
        """Return copy with paths resolved against root when relative."""
        resolved_paths = {
            name: (root / path if not path.is_absolute() else path)
            for name, path in self.paths.dict().items()
        }
        resolved_mappings = {
            name: (root / path if not path.is_absolute() else path)
            for name, path in self.mappings.aliases.items()
        }
        resolved_core = self.mappings.core
        if not resolved_core.is_absolute():
            resolved_core = root / resolved_core

        return AppConfig(
            paths=PathConfig(**resolved_paths),
            mappings=MappingConfig(core=resolved_core, aliases=resolved_mappings),
            processing=self.processing,
        )


def load_config(path: Path) -> AppConfig:
    data = yaml.safe_load(Path(path).read_text(encoding="utf-8"))
    config = AppConfig(**data)
    root = Path(path).resolve().parent
    return config.resolve(root)


def ensure_directories(paths: Iterable[Path]) -> None:
    for directory in paths:
        directory.mkdir(parents=True, exist_ok=True)


def list_platform_files(input_dir: Path, platform: str) -> List[Path]:
    platform_path = input_dir / platform
    if not platform_path.exists():
        logger.warning("Input directory for platform {platform} not found at {path}", platform=platform, path=platform_path)
        return []
    files = sorted(itertools.chain(platform_path.glob("*.xlsx"), platform_path.glob("*.csv")))
    logger.debug("Discovered {count} files for platform {platform}", count=len(files), platform=platform)
    return files


def _read_csv_with_detection(path: Path) -> pd.DataFrame:
    encodings = ["utf-8", "utf-8-sig", "cp1251", "cp866", "ISO-8859-1"]
    delimiters = [",", ";", "\t", "|"]
    last_error: Optional[Exception] = None
    for encoding in encodings:
        for delimiter in delimiters:
            try:
                df = pd.read_csv(path, encoding=encoding, sep=delimiter)
                if df.shape[1] <= 1 and delimiter != ",":
                    continue
                return df
            except Exception as exc:  # pragma: no cover - fallback branch
                last_error = exc
                continue
    if last_error:
        raise last_error
    raise ValueError(f"Unable to read CSV file {path}")


def read_input_table(path: Path) -> pd.DataFrame:
    logger.debug("Reading input file {path}", path=path)
    if path.suffix.lower() == ".csv":
        return _read_csv_with_detection(path)
    return pd.read_excel(path)


def read_yaml(path: Path) -> Mapping[str, List[str]]:
    with Path(path).open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    return {k: list(v) if isinstance(v, (list, tuple)) else [v] for k, v in data.items()}


def load_base_sheets(path: Path, required_sheets: Iterable[str]) -> Dict[str, pd.DataFrame]:
    file_path = Path(path)
    if not file_path.exists():
        logger.warning("Base workbook not found at {path}. A new file will be created.", path=path)
        return {sheet: pd.DataFrame() for sheet in required_sheets}

    wb = load_workbook(filename=file_path, read_only=False, data_only=True)
    sheets: Dict[str, pd.DataFrame] = {}
    for sheet_name in required_sheets:
        if sheet_name in wb.sheetnames:
            sheet_df = pd.read_excel(file_path, sheet_name=sheet_name)
            sheets[sheet_name] = sheet_df
        else:
            sheets[sheet_name] = pd.DataFrame()
    other_sheets = {name for name in wb.sheetnames if name not in required_sheets}
    for sheet_name in other_sheets:
        sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
    return sheets


def write_workbook(path: Path, sheets: Mapping[str, pd.DataFrame]) -> None:
    output_path = Path(path)
    ensure_directories([output_path.parent])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def prompt_save_path(default: Path) -> Path:
    response = typer.prompt("Куда сохранить итоговый файл?", default=str(default))
    return Path(response)


def timestamped_filename(prefix: str, suffix: str = ".xlsx") -> str:
    now = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{now}{suffix}"


def dataframe_to_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    return buffer.getvalue()


@contextlib.contextmanager
def excel_writer(path: Path):
    ensure_directories([path.parent])
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        yield writer
